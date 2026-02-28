#!/usr/bin/env python3
"""
Phase 1 multimodal extraction pipeline:
- PDF: text blocks + images (with nearest caption matching)
- Excel: sheet table extraction + embedded images
- Image processing: vision description (default), with optional OCR-first fallback mode
- Vector-ready output: JSONL chunks with consistent metadata
- Optional direct Chroma ingestion
"""

from __future__ import annotations

import argparse
import base64
import hashlib
import io
import json
import math
import os
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
import pytesseract
from PIL import Image

try:
    import fitz  # PyMuPDF
except Exception:
    fitz = None

try:
    from anthropic import Anthropic
except Exception:
    Anthropic = None

try:
    from openai import OpenAI
except Exception:
    OpenAI = None


SUPPORTED_INPUT_EXTENSIONS = {".pdf", ".xlsx"}


@dataclass
class OCRResult:
    text: str
    word_count: int
    avg_conf: float
    is_scanned_text: bool


@dataclass
class VisionProcessResult:
    structured: dict[str, Any] | None
    error: str | None
    calls_attempted: int
    calls_succeeded: int
    call_errors: int
    input_tokens: int
    output_tokens: int
    total_tokens: int
    retry_used: bool
    tiled: bool
    tile_count: int
    tile_rows: int
    tile_cols: int
    image_width: int
    image_height: int
    image_pixels: int
    json_invalid_after_retry: bool


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Phase 1 multimodal extraction pipeline for PDF and XLSX.")
    parser.add_argument("--data-dir", required=True, help="Root folder containing input files.")
    parser.add_argument("--output-dir", default="outputs/phase1_pipeline", help="Output directory.")
    parser.add_argument(
        "--vision-provider",
        default="openai",
        choices=["none", "anthropic", "openai"],
        help="Vision model provider for image descriptions.",
    )
    parser.add_argument(
        "--vision-model",
        default="gpt-4o",
        help="Vision model name for selected provider.",
    )
    parser.add_argument(
        "--vision-max-tokens",
        type=int,
        default=1800,
        help="First-attempt max tokens for image description generation.",
    )
    parser.add_argument(
        "--vision-retry-max-tokens",
        type=int,
        default=2500,
        help="Retry max tokens when first response is not valid JSON.",
    )
    parser.add_argument(
        "--vision-input-cost-per-1m",
        type=float,
        default=0.0,
        help="Optional cost estimate input rate (USD per 1M input tokens).",
    )
    parser.add_argument(
        "--vision-output-cost-per-1m",
        type=float,
        default=0.0,
        help="Optional cost estimate output rate (USD per 1M output tokens).",
    )
    parser.add_argument(
        "--ocr-word-threshold",
        type=int,
        default=45,
        help="OCR-first mode only: if OCR words >= threshold, image is treated as scanned text.",
    )
    parser.add_argument(
        "--ocr-char-threshold",
        type=int,
        default=280,
        help="OCR-first mode only: if OCR chars >= threshold, image is treated as scanned text.",
    )
    parser.add_argument(
        "--ocr-min-confidence-for-scanned",
        type=float,
        default=55.0,
        help=(
            "OCR-first mode only: minimum OCR confidence required before image is considered scanned text."
        ),
    )
    parser.add_argument(
        "--image-handling",
        default="vision_only",
        choices=["vision_only", "ocr_then_vision"],
        help=(
            "vision_only: all images go to vision model and are stored as image_description "
            "(default, matches Phase 1 spec). "
            "ocr_then_vision: classify scanned-text images to OCR, otherwise vision."
        ),
    )
    parser.add_argument(
        "--image-large-pixels-threshold",
        type=int,
        default=1_000_000,
        help="If width*height >= threshold, image is treated as large and tiled for vision.",
    )
    parser.add_argument(
        "--image-tile-target-max-pixels",
        type=int,
        default=1_000_000,
        help="Approximate max pixels per tile for large-image tiling path.",
    )
    parser.add_argument(
        "--image-max-tiles",
        type=int,
        default=9,
        help="Safety cap on number of tiles per image.",
    )
    parser.add_argument(
        "--vision-max-visible-text-items",
        type=int,
        default=120,
        help="Cap visible_text list length to keep JSON responses bounded.",
    )
    parser.add_argument("--max-pdf-pages", type=int, default=120, help="Max pages per PDF to process.")
    parser.add_argument("--max-sheet-rows", type=int, default=600, help="Max rows scanned per Excel sheet.")
    parser.add_argument("--max-sheet-cols", type=int, default=60, help="Max columns scanned per Excel sheet.")
    parser.add_argument(
        "--table-rows-per-chunk",
        type=int,
        default=35,
        help="Rows per table chunk in XLSX output.",
    )
    parser.add_argument(
        "--text-chunk-chars",
        type=int,
        default=1800,
        help="Approximate chars per text chunk.",
    )
    parser.add_argument(
        "--text-chunk-overlap",
        type=int,
        default=140,
        help="Overlap chars between adjacent text chunks.",
    )
    parser.add_argument(
        "--vector-db",
        default="none",
        choices=["none", "chroma"],
        help="Optional direct vector DB ingestion.",
    )
    parser.add_argument(
        "--chroma-path",
        default="outputs/phase1_pipeline/chroma_db",
        help="Chroma persistence directory when --vector-db chroma.",
    )
    parser.add_argument(
        "--chroma-collection",
        default="phase1_chunks",
        help="Chroma collection name.",
    )
    parser.add_argument(
        "--chroma-batch-size",
        type=int,
        default=100,
        help="Batch size for Chroma insertion.",
    )
    return parser.parse_args()


def now_utc_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def clean_text(text: str) -> str:
    text = text.replace("\x00", " ")
    text = text.replace("\u2014", "-")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def to_string_list(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, str):
        v = clean_text(value)
        return [v] if v else []
    if isinstance(value, list):
        out: list[str] = []
        for item in value:
            s = clean_text(str(item))
            if s:
                out.append(s)
        return out
    s = clean_text(str(value))
    return [s] if s else []


def extract_json_object(text: str) -> dict[str, Any] | None:
    text = text.strip()
    if not text:
        return None

    # Best effort: handle markdown-fenced JSON blocks.
    fenced = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, flags=re.IGNORECASE | re.DOTALL)
    if fenced:
        candidate = fenced.group(1).strip()
        try:
            parsed = json.loads(candidate)
            if isinstance(parsed, dict):
                return parsed
        except Exception:
            pass

    try:
        parsed = json.loads(text)
        if isinstance(parsed, dict):
            return parsed
    except Exception:
        pass

    start = text.find("{")
    while start != -1:
        depth = 0
        for i in range(start, len(text)):
            ch = text[i]
            if ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    candidate = text[start : i + 1]
                    try:
                        parsed = json.loads(candidate)
                        if isinstance(parsed, dict):
                            return parsed
                    except Exception:
                        break
        start = text.find("{", start + 1)
    return None


def normalize_vision_output(obj: dict[str, Any] | None, raw_text: str) -> dict[str, Any]:
    if obj is None:
        return {
            "image_type": "unknown",
            "visible_text": [],
            "description": clean_text(raw_text) or "[no_description_returned]",
            "elements": [],
            "layout": "",
            "unclear_parts": "",
            "json_parse_fallback_used": True,
        }
    image_type = clean_text(str(obj.get("image_type", "unknown"))) or "unknown"
    visible_text = to_string_list(obj.get("visible_text", []))
    description = clean_text(str(obj.get("description", "")))
    elements = to_string_list(obj.get("elements", []))
    layout = clean_text(str(obj.get("layout", "")))
    unclear_parts = clean_text(str(obj.get("unclear_parts", "")))
    if not description:
        description = "[no_description_returned]"
    return {
        "image_type": image_type,
        "visible_text": visible_text,
        "description": description,
        "elements": elements,
        "layout": layout,
        "unclear_parts": unclear_parts,
        "json_parse_fallback_used": False,
    }


def build_image_text_content(vision_struct: dict[str, Any]) -> str:
    image_type = clean_text(str(vision_struct.get("image_type", "unknown"))) or "unknown"
    visible_text_list = to_string_list(vision_struct.get("visible_text", []))
    description = clean_text(str(vision_struct.get("description", ""))) or "[no_description_returned]"
    elements = to_string_list(vision_struct.get("elements", []))
    layout = clean_text(str(vision_struct.get("layout", "")))
    unclear_parts = clean_text(str(vision_struct.get("unclear_parts", "")))

    sections: list[str] = []
    sections.append(f"Image type: {image_type}")
    sections.append(
        "Extracted visible text:\n"
        + ("\n".join(f"- {line}" for line in visible_text_list) if visible_text_list else "- [none]")
    )
    sections.append(f"Description: {description}")
    if elements:
        sections.append("Elements:\n" + "\n".join(f"- {item}" for item in elements))
    if layout:
        sections.append(f"Layout: {layout}")
    if unclear_parts:
        sections.append(f"Unclear parts: {unclear_parts}")
    return "\n\n".join(sections).strip()


def dedupe_preserve_order(values: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for v in values:
        s = clean_text(v)
        if not s or s in seen:
            continue
        seen.add(s)
        out.append(s)
    return out


def choose_merged_image_type(image_types: list[str], tiled: bool) -> str:
    cleaned = [clean_text(v) for v in image_types if clean_text(v) and clean_text(v) != "unknown"]
    if not cleaned:
        return "tiled_image" if tiled else "unknown"
    counts: dict[str, int] = {}
    for t in cleaned:
        counts[t] = counts.get(t, 0) + 1
    top = sorted(counts.items(), key=lambda kv: (-kv[1], kv[0]))[0][0]
    return top


def merge_vision_tile_outputs(tile_outputs: list[dict[str, Any]], tiled: bool) -> dict[str, Any]:
    image_types = [str(t.get("image_type", "unknown")) for t in tile_outputs]
    visible_text: list[str] = []
    elements: list[str] = []
    description_parts: list[str] = []
    layout_parts: list[str] = []
    unclear_parts_list: list[str] = []
    parse_fallback_used = False

    for i, t in enumerate(tile_outputs, 1):
        v_lines = to_string_list(t.get("visible_text", []))
        visible_text.extend(v_lines)
        elements.extend(to_string_list(t.get("elements", [])))

        desc = clean_text(str(t.get("description", "")))
        if desc:
            if tiled:
                description_parts.append(f"Tile {i}: {desc}")
            else:
                description_parts.append(desc)

        layout = clean_text(str(t.get("layout", "")))
        if layout:
            if tiled:
                layout_parts.append(f"Tile {i}: {layout}")
            else:
                layout_parts.append(layout)

        unclear = clean_text(str(t.get("unclear_parts", "")))
        if unclear:
            unclear_parts_list.append(unclear)

        if bool(t.get("json_parse_fallback_used", False)):
            parse_fallback_used = True

    merged = {
        "image_type": choose_merged_image_type(image_types, tiled=tiled),
        "visible_text": dedupe_preserve_order(visible_text),
        "description": clean_text(" ".join(description_parts)) or "[no_description_returned]",
        "elements": dedupe_preserve_order(elements),
        "layout": clean_text(" ".join(layout_parts)),
        "unclear_parts": clean_text(" | ".join(dedupe_preserve_order(unclear_parts_list))),
        "json_parse_fallback_used": parse_fallback_used,
    }
    return merged


def compute_tile_grid(width: int, height: int, target_max_pixels: int, max_tiles: int) -> tuple[int, int]:
    target = max(1, int(target_max_pixels))
    max_tiles = max(2, int(max_tiles))
    total_pixels = max(1, width * height)
    desired_tiles = max(2, math.ceil(total_pixels / target))
    desired_tiles = min(desired_tiles, max_tiles)

    best_rows = 1
    best_cols = desired_tiles
    best_score = float("inf")

    for rows in range(1, desired_tiles + 1):
        cols = math.ceil(desired_tiles / rows)
        if rows * cols > max_tiles:
            continue
        tile_w = math.ceil(width / cols)
        tile_h = math.ceil(height / rows)
        tile_pixels = tile_w * tile_h
        aspect_penalty = abs((tile_w / max(1, tile_h)) - 1.0)
        score = tile_pixels + aspect_penalty * 10000 + (rows * cols - desired_tiles) * 100
        if score < best_score:
            best_score = score
            best_rows, best_cols = rows, cols

    if best_rows * best_cols < 2:
        if width >= height:
            best_rows, best_cols = 1, 2
        else:
            best_rows, best_cols = 2, 1
    return best_rows, best_cols


def split_image_into_tiles(
    image_bytes: bytes,
    target_max_pixels: int,
    max_tiles: int,
) -> tuple[list[dict[str, Any]], int, int, int, int, int]:
    with Image.open(io.BytesIO(image_bytes)) as img:
        rgb = img.convert("RGB")
        width, height = rgb.size
        total_pixels = width * height
        rows, cols = compute_tile_grid(width, height, target_max_pixels, max_tiles)

        tile_w = math.ceil(width / cols)
        tile_h = math.ceil(height / rows)

        tiles: list[dict[str, Any]] = []
        idx = 0
        for r in range(rows):
            for c in range(cols):
                left = c * tile_w
                upper = r * tile_h
                right = min(width, left + tile_w)
                lower = min(height, upper + tile_h)
                if right <= left or lower <= upper:
                    continue
                crop = rgb.crop((left, upper, right, lower))
                buf = io.BytesIO()
                crop.save(buf, format="PNG")
                idx += 1
                tiles.append(
                    {
                        "index": idx,
                        "row": r + 1,
                        "col": c + 1,
                        "bbox": {"x0": left, "y0": upper, "x1": right, "y1": lower},
                        "width": right - left,
                        "height": lower - upper,
                        "pixels": (right - left) * (lower - upper),
                        "bytes": buf.getvalue(),
                        "mime_type": "image/png",
                    }
                )
        return tiles, width, height, total_pixels, rows, cols


def safe_mkdir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def sha1_id(raw: str) -> str:
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()


def chunk_text(text: str, max_chars: int, overlap: int) -> list[str]:
    text = clean_text(text)
    if not text:
        return []
    if len(text) <= max_chars:
        return [text]

    chunks: list[str] = []
    start = 0
    n = len(text)
    while start < n:
        end = min(n, start + max_chars)
        if end < n:
            cut = text.rfind(" ", start + int(max_chars * 0.6), end)
            if cut > start:
                end = cut
        piece = text[start:end].strip()
        if piece:
            chunks.append(piece)
        if end >= n:
            break
        start = max(0, end - overlap)
    return chunks


def list_target_files(data_dir: Path) -> list[Path]:
    out: list[Path] = []
    for p in sorted(data_dir.rglob("*")):
        if p.is_file() and p.suffix.lower() in SUPPORTED_INPUT_EXTENSIONS:
            out.append(p)
    return out


def guess_mime(ext: str) -> str:
    e = ext.lower().lstrip(".")
    if e in {"jpg", "jpeg"}:
        return "image/jpeg"
    if e == "png":
        return "image/png"
    if e == "gif":
        return "image/gif"
    if e in {"tif", "tiff"}:
        return "image/tiff"
    if e == "bmp":
        return "image/bmp"
    return "application/octet-stream"


class VisionDescriber:
    def __init__(self, provider: str, model: str, max_tokens: int) -> None:
        self.provider = provider
        self.model = model
        self.max_tokens = max_tokens
        self._client = None
        if provider == "anthropic":
            if Anthropic is None:
                raise RuntimeError("anthropic package is not installed.")
            api_key = os.getenv("ANTHROPIC_API_KEY")
            if not api_key:
                raise RuntimeError("ANTHROPIC_API_KEY is not set.")
            self._client = Anthropic(api_key=api_key)
        elif provider == "openai":
            if OpenAI is None:
                raise RuntimeError("openai package is not installed.")
            api_key = os.getenv("OPENAI_API_KEY")
            if not api_key:
                raise RuntimeError("OPENAI_API_KEY is not set.")
            self._client = OpenAI(api_key=api_key)

    def describe(
        self,
        image_bytes: bytes,
        mime_type: str,
        caption_hint: str,
        source_file: str,
        locator: str,
        max_tokens: int | None = None,
        retry_note: str | None = None,
        max_visible_text_items: int = 120,
    ) -> tuple[dict[str, Any], dict[str, int]]:
        effective_max_tokens = int(max_tokens or self.max_tokens)
        prompt = (
            "You are analyzing an academic submission image for grading support.\n"
            "Rules:\n"
            "- Only describe what is literally visible in the image.\n"
            "- Do not infer, assume, or add information not visible.\n"
            "- If text or labels are unreadable, include them under unclear_parts.\n"
            "- Extract visible text exactly as it appears where possible.\n"
            f"- Include at most {int(max_visible_text_items)} entries in visible_text.\n"
            "Return ONLY valid JSON with this exact schema and keys:\n"
            "{\n"
            '  "image_type": "string",\n'
            '  "visible_text": ["string"],\n'
            '  "description": "string",\n'
            '  "elements": ["string"],\n'
            '  "layout": "string",\n'
            '  "unclear_parts": "string"\n'
            "}\n"
            f"Source: {source_file} ({locator}).\n"
            f"Caption hint (may be empty): {caption_hint or '[none]'}"
        )
        if retry_note:
            prompt += f"\nRetry instruction: {retry_note}"

        if self.provider == "anthropic":
            assert self._client is not None
            b64 = base64.b64encode(image_bytes).decode("ascii")
            message = self._client.messages.create(
                model=self.model,
                max_tokens=effective_max_tokens,
                temperature=0,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": prompt},
                            {
                                "type": "image",
                                "source": {
                                    "type": "base64",
                                    "media_type": mime_type,
                                    "data": b64,
                                },
                            },
                        ],
                    }
                ],
            )
            parts = []
            for part in message.content:
                if getattr(part, "type", None) == "text":
                    parts.append(getattr(part, "text", ""))
            raw_text = clean_text(" ".join(parts))
            parsed_obj = extract_json_object(raw_text)
            normalized = normalize_vision_output(parsed_obj, raw_text)
            usage_obj = getattr(message, "usage", None)
            input_tokens = int(getattr(usage_obj, "input_tokens", 0) or 0)
            output_tokens = int(getattr(usage_obj, "output_tokens", 0) or 0)
            total_tokens = input_tokens + output_tokens
            return normalized, {
                "input_tokens": input_tokens,
                "output_tokens": output_tokens,
                "total_tokens": total_tokens,
            }

        if self.provider == "openai":
            assert self._client is not None
            b64 = base64.b64encode(image_bytes).decode("ascii")
            data_url = f"data:{mime_type};base64,{b64}"
            payload = {
                "model": self.model,
                "input": [
                    {
                        "role": "user",
                        "content": [
                            {"type": "input_text", "text": prompt},
                            {"type": "input_image", "image_url": data_url},
                        ],
                    }
                ],
                "max_output_tokens": effective_max_tokens,
                "temperature": 0,
            }
            try:
                response = self._client.responses.create(**payload)
            except TypeError:
                payload.pop("temperature", None)
                response = self._client.responses.create(**payload)
            raw_text = clean_text(getattr(response, "output_text", "") or "")
            parsed_obj = extract_json_object(raw_text)
            normalized = normalize_vision_output(parsed_obj, raw_text)
            usage_obj = getattr(response, "usage", None)
            input_tokens = int(getattr(usage_obj, "input_tokens", 0) or 0)
            output_tokens = int(getattr(usage_obj, "output_tokens", 0) or 0)
            total_tokens = int(getattr(usage_obj, "total_tokens", 0) or (input_tokens + output_tokens))
            return normalized, {
                "input_tokens": input_tokens,
                "output_tokens": output_tokens,
                "total_tokens": total_tokens,
            }

        raise RuntimeError(f"Unsupported vision provider: {self.provider}")


def describe_image_with_strategy(
    vision: VisionDescriber,
    image_bytes: bytes,
    mime_type: str,
    caption_hint: str,
    source_file: str,
    locator: str,
    args: argparse.Namespace,
) -> VisionProcessResult:
    with Image.open(io.BytesIO(image_bytes)) as img:
        width, height = img.size
    image_pixels = int(width * height)

    large_image = image_pixels >= int(args.image_large_pixels_threshold)
    if large_image:
        tiles, width, height, image_pixels, tile_rows, tile_cols = split_image_into_tiles(
            image_bytes=image_bytes,
            target_max_pixels=int(args.image_tile_target_max_pixels),
            max_tiles=int(args.image_max_tiles),
        )
    else:
        tile_rows, tile_cols = 1, 1
        tiles = [
            {
                "index": 1,
                "row": 1,
                "col": 1,
                "bbox": {"x0": 0, "y0": 0, "x1": width, "y1": height},
                "width": width,
                "height": height,
                "pixels": image_pixels,
                "bytes": image_bytes,
                "mime_type": mime_type,
            }
        ]

    calls_attempted = 0
    calls_succeeded = 0
    call_errors = 0
    input_tokens = 0
    output_tokens = 0
    total_tokens = 0
    retry_used = False
    json_invalid_after_retry = False
    tile_outputs: list[dict[str, Any]] = []

    for tile in tiles:
        tile_locator = locator
        if large_image:
            tile_locator = (
                f"{locator}, tile {tile['index']}/{len(tiles)} "
                f"(row {tile['row']}, col {tile['col']})"
            )

        calls_attempted += 1
        try:
            first_structured, first_usage = vision.describe(
                image_bytes=tile["bytes"],
                mime_type=tile["mime_type"],
                caption_hint=caption_hint,
                source_file=source_file,
                locator=tile_locator,
                max_tokens=int(args.vision_max_tokens),
                max_visible_text_items=int(args.vision_max_visible_text_items),
            )
            calls_succeeded += 1
            input_tokens += int(first_usage.get("input_tokens", 0))
            output_tokens += int(first_usage.get("output_tokens", 0))
            total_tokens += int(first_usage.get("total_tokens", 0))
        except Exception as exc:
            call_errors += 1
            return VisionProcessResult(
                structured=None,
                error=f"vision_call_failed ({tile_locator}): {exc}",
                calls_attempted=calls_attempted,
                calls_succeeded=calls_succeeded,
                call_errors=call_errors,
                input_tokens=input_tokens,
                output_tokens=output_tokens,
                total_tokens=total_tokens,
                retry_used=retry_used,
                tiled=large_image,
                tile_count=len(tiles),
                tile_rows=tile_rows,
                tile_cols=tile_cols,
                image_width=width,
                image_height=height,
                image_pixels=image_pixels,
                json_invalid_after_retry=json_invalid_after_retry,
            )

        if not bool(first_structured.get("json_parse_fallback_used", False)):
            tile_outputs.append(first_structured)
            continue

        retry_tokens = int(args.vision_retry_max_tokens)
        if retry_tokens <= int(args.vision_max_tokens):
            json_invalid_after_retry = True
            tile_outputs.append(first_structured)
            continue

        retry_used = True
        calls_attempted += 1
        try:
            second_structured, second_usage = vision.describe(
                image_bytes=tile["bytes"],
                mime_type=tile["mime_type"],
                caption_hint=caption_hint,
                source_file=source_file,
                locator=tile_locator,
                max_tokens=retry_tokens,
                retry_note=(
                    "Previous response was not valid JSON. Return only valid JSON matching "
                    "the required schema and keys."
                ),
                max_visible_text_items=int(args.vision_max_visible_text_items),
            )
            calls_succeeded += 1
            input_tokens += int(second_usage.get("input_tokens", 0))
            output_tokens += int(second_usage.get("output_tokens", 0))
            total_tokens += int(second_usage.get("total_tokens", 0))
            tile_outputs.append(second_structured)
            if bool(second_structured.get("json_parse_fallback_used", False)):
                json_invalid_after_retry = True
        except Exception:
            call_errors += 1
            json_invalid_after_retry = True
            # Keep first response as fallback content for this tile.
            tile_outputs.append(first_structured)

    if not tile_outputs:
        return VisionProcessResult(
            structured=None,
            error="vision_strategy_failed_without_output",
            calls_attempted=calls_attempted,
            calls_succeeded=calls_succeeded,
            call_errors=call_errors,
            input_tokens=input_tokens,
            output_tokens=output_tokens,
            total_tokens=total_tokens,
            retry_used=retry_used,
            tiled=large_image,
            tile_count=len(tiles),
            tile_rows=tile_rows,
            tile_cols=tile_cols,
            image_width=width,
            image_height=height,
            image_pixels=image_pixels,
            json_invalid_after_retry=json_invalid_after_retry,
        )

    merged = merge_vision_tile_outputs(tile_outputs, tiled=large_image)
    if json_invalid_after_retry:
        merged["json_parse_fallback_used"] = True

    return VisionProcessResult(
        structured=merged,
        error=None,
        calls_attempted=calls_attempted,
        calls_succeeded=calls_succeeded,
        call_errors=call_errors,
        input_tokens=input_tokens,
        output_tokens=output_tokens,
        total_tokens=total_tokens,
        retry_used=retry_used,
        tiled=large_image,
        tile_count=len(tiles),
        tile_rows=tile_rows,
        tile_cols=tile_cols,
        image_width=width,
        image_height=height,
        image_pixels=image_pixels,
        json_invalid_after_retry=json_invalid_after_retry,
    )


def compute_ocr(image_bytes: bytes) -> OCRResult:
    with Image.open(io.BytesIO(image_bytes)) as img:
        img = img.convert("RGB")
        data = pytesseract.image_to_data(img, output_type=pytesseract.Output.DICT)
        words: list[str] = []
        confs: list[float] = []
        for i in range(len(data["text"])):
            t = clean_text(data["text"][i])
            if not t:
                continue
            words.append(t)
            try:
                c = float(data["conf"][i])
            except Exception:
                c = -1.0
            if c >= 0:
                confs.append(c)
        text = clean_text(" ".join(words))
        avg_conf = round(sum(confs) / len(confs), 2) if confs else 0.0
        return OCRResult(
            text=text,
            word_count=len(words),
            avg_conf=avg_conf,
            is_scanned_text=False,
        )


def bbox_center(b: dict[str, float]) -> tuple[float, float]:
    return ((b["x0"] + b["x1"]) / 2.0, (b["y0"] + b["y1"]) / 2.0)


def bbox_overlap_x(a: dict[str, float], b: dict[str, float]) -> float:
    left = max(a["x0"], b["x0"])
    right = min(a["x1"], b["x1"])
    if right <= left:
        return 0.0
    overlap = right - left
    aw = max(1.0, a["x1"] - a["x0"])
    bw = max(1.0, b["x1"] - b["x0"])
    return overlap / min(aw, bw)


def match_nearest_caption(
    image_bbox: dict[str, float],
    text_blocks: list[dict[str, Any]],
) -> tuple[str, str | None, float | None]:
    if not text_blocks:
        return "", None, None

    candidates: list[tuple[float, dict[str, Any], str]] = []
    ix0, iy0, ix1, iy1 = image_bbox["x0"], image_bbox["y0"], image_bbox["x1"], image_bbox["y1"]
    icx, icy = bbox_center(image_bbox)

    for block in text_blocks:
        b = block["bbox"]
        bcx, bcy = bbox_center(b)
        overlap_x = bbox_overlap_x(image_bbox, b)
        vertical_gap_below = b["y0"] - iy1
        vertical_gap_above = iy0 - b["y1"]

        if vertical_gap_below >= 0:
            # Prefer captions below image.
            distance = vertical_gap_below + abs(bcx - icx) * 0.15
            if overlap_x > 0.05 or vertical_gap_below < 180:
                candidates.append((distance, block, "below"))
        elif vertical_gap_above >= 0:
            # Fallback: text above image.
            distance = vertical_gap_above + abs(bcx - icx) * 0.2 + 25
            if overlap_x > 0.05 or vertical_gap_above < 130:
                candidates.append((distance, block, "above"))
        else:
            # Overlapping text block.
            dist = math.dist((icx, icy), (bcx, bcy))
            candidates.append((dist + 40, block, "overlap"))

    if not candidates:
        # Pure nearest fallback.
        nearest = min(
            text_blocks,
            key=lambda block: math.dist(bbox_center(image_bbox), bbox_center(block["bbox"])),
        )
        dist = math.dist(bbox_center(image_bbox), bbox_center(nearest["bbox"]))
        return nearest["text"], nearest["id"], dist

    best = min(candidates, key=lambda x: x[0])
    return best[1]["text"], best[1]["id"], round(best[0], 2)


def pdf_text_blocks(page: Any) -> list[dict[str, Any]]:
    blocks = page.get_text("blocks")
    out: list[dict[str, Any]] = []
    for i, block in enumerate(blocks):
        if len(block) < 6:
            continue
        x0, y0, x1, y1, text = block[:5]
        btype = block[6] if len(block) >= 7 else 0
        if btype != 0:
            continue
        text = clean_text(str(text))
        if not text:
            continue
        out.append(
            {
                "id": f"T{i+1}",
                "bbox": {"x0": float(x0), "y0": float(y0), "x1": float(x1), "y1": float(y1)},
                "text": text,
            }
        )
    out.sort(key=lambda t: (t["bbox"]["y0"], t["bbox"]["x0"]))
    return out


def save_image(image_bytes: bytes, image_path: Path) -> None:
    safe_mkdir(image_path.parent)
    image_path.write_bytes(image_bytes)


def process_pdf(
    file_path: Path,
    rel_path: str,
    images_dir: Path,
    args: argparse.Namespace,
    vision: VisionDescriber | None,
) -> tuple[list[dict[str, Any]], dict[str, int]]:
    if fitz is None:
        raise RuntimeError("PyMuPDF is not installed. Add PyMuPDF to requirements and install it.")

    chunks: list[dict[str, Any]] = []
    stats = {
        "text_chunks": 0,
        "image_chunks": 0,
        "ocr_chunks": 0,
        "vision_chunks": 0,
        "vision_errors": 0,
        "vision_json_fallbacks": 0,
        "vision_retries": 0,
        "tiled_images": 0,
        "images_routed_to_vision": 0,
        "vision_calls_attempted": 0,
        "vision_calls_succeeded": 0,
        "vision_input_tokens": 0,
        "vision_output_tokens": 0,
        "vision_total_tokens": 0,
    }

    doc = fitz.open(str(file_path))
    try:
        page_limit = min(len(doc), args.max_pdf_pages)
        for pidx in range(page_limit):
            page = doc[pidx]
            page_num = pidx + 1
            text_blocks = pdf_text_blocks(page)

            # Text chunks
            for block in text_blocks:
                for ci, chunk in enumerate(
                    chunk_text(block["text"], args.text_chunk_chars, args.text_chunk_overlap),
                    1,
                ):
                    # Skip page-number fragments and tiny punctuation chunks.
                    if len(chunk) < 8 and re.fullmatch(r"[\d\W_]+", chunk):
                        continue
                    meta = {
                        "filename": file_path.name,
                        "source_path": rel_path,
                        "page": page_num,
                        "content_type": "text",
                        "block_id": block["id"],
                        "bbox": block["bbox"],
                        "chunk_index_in_block": ci,
                    }
                    cid = sha1_id(f"{rel_path}|page={page_num}|text|{block['id']}|{ci}")
                    chunks.append({"id": cid, "content": chunk, "metadata": meta})
                    stats["text_chunks"] += 1

            # Image chunks
            page_images = page.get_images(full=True)
            for img_i, img_info in enumerate(page_images, 1):
                xref = img_info[0]
                img_data = doc.extract_image(xref)
                if not img_data or "image" not in img_data:
                    continue

                image_bytes = img_data["image"]
                ext = str(img_data.get("ext", "png")).lower()
                image_path = images_dir / rel_path / f"page_{page_num:03d}_img_{img_i:03d}.{ext}"
                save_image(image_bytes, image_path)

                rects = page.get_image_rects(xref)
                rect = rects[0] if rects else None
                bbox = {
                    "x0": float(rect.x0) if rect else None,
                    "y0": float(rect.y0) if rect else None,
                    "x1": float(rect.x1) if rect else None,
                    "y1": float(rect.y1) if rect else None,
                }

                caption_text, caption_block_id, caption_distance = (
                    ("", None, None) if rect is None else match_nearest_caption(bbox, text_blocks)
                )

                ocr = compute_ocr(image_bytes)
                scanned_by_volume = (
                    len(ocr.text) >= args.ocr_char_threshold or ocr.word_count >= args.ocr_word_threshold
                )
                scanned = scanned_by_volume and (ocr.avg_conf >= args.ocr_min_confidence_for_scanned)
                force_vision = args.image_handling == "vision_only"
                should_use_vision = force_vision or (not scanned)

                content_type = "ocr_text"
                content = ocr.text
                vision_used = False
                vision_error = None
                vision_structured: dict[str, Any] | None = None
                tiled = False
                tile_count = 1
                tile_rows = 1
                tile_cols = 1
                image_width = None
                image_height = None
                image_pixels = None
                vision_retry_used = False
                vision_json_invalid_after_retry = False

                if should_use_vision:
                    stats["images_routed_to_vision"] += 1
                    if vision is None:
                        vision_error = "vision_client_not_ready"
                        if force_vision:
                            fallback = []
                            fallback.append(f"[vision_error] {vision_error}")
                            if caption_text:
                                fallback.append(f"Caption hint: {caption_text}")
                            if ocr.text:
                                fallback.append(f"OCR fallback: {ocr.text}")
                            content = clean_text(" ".join(fallback)) or "[vision_error_no_fallback_text]"
                            content_type = "image_description"
                        else:
                            fallback = []
                            if caption_text:
                                fallback.append(f"Caption hint: {caption_text}")
                            if ocr.text:
                                fallback.append(f"OCR text: {ocr.text}")
                            content = clean_text(" ".join(fallback)) or "[vision_failed_and_no_ocr_text]"
                            content_type = "ocr_text"
                    else:
                        vision_result = describe_image_with_strategy(
                            vision=vision,
                            image_bytes=image_bytes,
                            mime_type=guess_mime(ext),
                            caption_hint=caption_text,
                            source_file=rel_path,
                            locator=f"page {page_num}, image {img_i}",
                            args=args,
                        )
                        stats["vision_calls_attempted"] += int(vision_result.calls_attempted)
                        stats["vision_calls_succeeded"] += int(vision_result.calls_succeeded)
                        stats["vision_errors"] += int(vision_result.call_errors)
                        stats["vision_input_tokens"] += int(vision_result.input_tokens)
                        stats["vision_output_tokens"] += int(vision_result.output_tokens)
                        stats["vision_total_tokens"] += int(vision_result.total_tokens)

                        tiled = bool(vision_result.tiled)
                        tile_count = int(vision_result.tile_count)
                        tile_rows = int(vision_result.tile_rows)
                        tile_cols = int(vision_result.tile_cols)
                        image_width = int(vision_result.image_width)
                        image_height = int(vision_result.image_height)
                        image_pixels = int(vision_result.image_pixels)
                        vision_retry_used = bool(vision_result.retry_used)
                        vision_json_invalid_after_retry = bool(vision_result.json_invalid_after_retry)
                        if tiled:
                            stats["tiled_images"] += 1
                        if vision_retry_used:
                            stats["vision_retries"] += 1

                        if vision_result.structured is not None:
                            vision_structured = vision_result.structured
                            content = build_image_text_content(vision_structured)
                            content_type = "image_description"
                            vision_used = True
                            if bool(vision_structured.get("json_parse_fallback_used", False)):
                                stats["vision_json_fallbacks"] += 1
                            if vision_result.error:
                                vision_error = vision_result.error
                        else:
                            vision_error = vision_result.error or "vision_strategy_failed_without_output"
                            # Keep content type as image_description in strict mode.
                            if force_vision:
                                fallback = []
                                fallback.append(f"[vision_error] {vision_error}")
                                if caption_text:
                                    fallback.append(f"Caption hint: {caption_text}")
                                if ocr.text:
                                    fallback.append(f"OCR fallback: {ocr.text}")
                                content = clean_text(" ".join(fallback)) or "[vision_error_no_fallback_text]"
                                content_type = "image_description"
                            else:
                                fallback = []
                                if caption_text:
                                    fallback.append(f"Caption hint: {caption_text}")
                                if ocr.text:
                                    fallback.append(f"OCR text: {ocr.text}")
                                content = clean_text(" ".join(fallback)) or "[vision_failed_and_no_ocr_text]"
                                content_type = "ocr_text"

                meta = {
                    "filename": file_path.name,
                    "source_path": rel_path,
                    "page": page_num,
                    "content_type": content_type,
                    "image_index_on_page": img_i,
                    "image_path": str(image_path),
                    "image_ext": ext,
                    "bbox": bbox,
                    "caption_text": caption_text,
                    "caption_block_id": caption_block_id,
                    "caption_distance": caption_distance,
                    "ocr_word_count": ocr.word_count,
                    "ocr_avg_conf": ocr.avg_conf,
                    "ocr_char_count": len(ocr.text),
                    "scanned_detected_by_volume": scanned_by_volume,
                    "scanned_text_detected": scanned,
                    "image_handling": args.image_handling,
                    "vision_used": vision_used,
                    "vision_provider": args.vision_provider if should_use_vision else None,
                    "vision_model": args.vision_model if should_use_vision else None,
                    "vision_error": vision_error,
                    "tiled": tiled,
                    "tile_count": tile_count,
                    "tile_rows": tile_rows,
                    "tile_cols": tile_cols,
                    "image_width": image_width,
                    "image_height": image_height,
                    "image_pixels": image_pixels,
                    "vision_retry_used": vision_retry_used,
                    "vision_json_invalid_after_retry": vision_json_invalid_after_retry,
                }
                if vision_structured is not None:
                    extracted_text_lines = to_string_list(vision_structured.get("visible_text", []))
                    meta["image_type"] = vision_structured.get("image_type")
                    meta["extracted_text"] = "\n".join(extracted_text_lines)
                    meta["visible_text_lines"] = extracted_text_lines
                    meta["description"] = vision_structured.get("description")
                    meta["elements"] = to_string_list(vision_structured.get("elements", []))
                    meta["layout"] = vision_structured.get("layout")
                    meta["unclear_parts"] = vision_structured.get("unclear_parts")
                    meta["vision_json_parse_fallback_used"] = bool(
                        vision_structured.get("json_parse_fallback_used", False)
                    )
                cid = sha1_id(f"{rel_path}|page={page_num}|image|{img_i}|{content_type}")
                chunks.append({"id": cid, "content": content, "metadata": meta})
                stats["image_chunks"] += 1
                if content_type == "ocr_text":
                    stats["ocr_chunks"] += 1
                if content_type == "image_description":
                    stats["vision_chunks"] += 1

    finally:
        doc.close()

    return chunks, stats


def find_header_row(sheet_rows: list[list[str]]) -> int | None:
    best_idx = None
    best_score = float("-inf")
    header_keywords = {
        "header",
        "function",
        "importance",
        "score",
        "notes",
        "description",
        "id",
        "name",
        "requirement",
        "vendor",
    }

    for i, row in enumerate(sheet_rows[:120]):
        non_empty = [v for v in row if v]
        n = len(non_empty)
        if n < 2:
            continue

        lengths = [len(v) for v in non_empty]
        avg_len = sum(lengths) / max(1, n)
        alpha_cells = sum(1 for v in non_empty if re.search(r"[A-Za-z]", v))
        keyword_hits = sum(
            1
            for v in non_empty
            for token in re.findall(r"[A-Za-z]+", v.lower())
            if token in header_keywords
        )

        score = 0.0
        score += n * 1.4
        score += alpha_cells * 1.0
        score += keyword_hits * 2.5
        # Header rows are usually short labels, not long narrative text.
        if 3 <= avg_len <= 30:
            score += 5.0
        if avg_len > 70:
            score -= 8.0
        if any(len(v) > 130 for v in non_empty):
            score -= 4.0
        if n >= 3:
            score += 2.0
        # Slight bias toward earlier rows.
        score -= i * 0.03

        if score > best_score:
            best_score = score
            best_idx = i

    return best_idx


def process_excel(
    file_path: Path,
    rel_path: str,
    images_dir: Path,
    args: argparse.Namespace,
    vision: VisionDescriber | None,
) -> tuple[list[dict[str, Any]], dict[str, int]]:
    chunks: list[dict[str, Any]] = []
    stats = {
        "table_chunks": 0,
        "image_chunks": 0,
        "ocr_chunks": 0,
        "vision_chunks": 0,
        "vision_errors": 0,
        "vision_json_fallbacks": 0,
        "vision_retries": 0,
        "tiled_images": 0,
        "images_routed_to_vision": 0,
        "vision_calls_attempted": 0,
        "vision_calls_succeeded": 0,
        "vision_input_tokens": 0,
        "vision_output_tokens": 0,
        "vision_total_tokens": 0,
    }

    wb = openpyxl.load_workbook(file_path, data_only=False)
    try:
        for ws in wb.worksheets:
            max_rows = min(ws.max_row or 1, args.max_sheet_rows)
            max_cols = min(ws.max_column or 1, args.max_sheet_cols)

            grid: list[list[str]] = []
            for r in ws.iter_rows(min_row=1, max_row=max_rows, min_col=1, max_col=max_cols, values_only=True):
                row_vals = [clean_text("" if v is None else str(v)) for v in r]
                grid.append(row_vals)

            header_idx = find_header_row(grid)
            header = grid[header_idx] if header_idx is not None else []

            data_lines: list[str] = []
            if header:
                header_line = " | ".join([h if h else f"col_{i+1}" for i, h in enumerate(header)])
            else:
                header_line = " | ".join([f"col_{i+1}" for i in range(max_cols)])

            start_row = (header_idx + 1) if header_idx is not None else 0
            for ridx in range(start_row, len(grid)):
                row = grid[ridx]
                if not any(row):
                    continue
                pairs = []
                for cidx, value in enumerate(row):
                    if not value:
                        continue
                    key = header[cidx] if cidx < len(header) and header[cidx] else f"col_{cidx+1}"
                    pairs.append(f"{key}: {value}")
                if pairs:
                    data_lines.append(f"row_{ridx+1}: " + " ; ".join(pairs))

            if data_lines:
                for chunk_start in range(0, len(data_lines), args.table_rows_per_chunk):
                    section = data_lines[chunk_start : chunk_start + args.table_rows_per_chunk]
                    text = (
                        f"Sheet: {ws.title}\n"
                        f"Headers: {header_line}\n"
                        "Rows:\n" + "\n".join(section)
                    )
                    meta = {
                        "filename": file_path.name,
                        "source_path": rel_path,
                        "sheet_name": ws.title,
                        "content_type": "table",
                        "row_start": chunk_start + 1,
                        "row_end": chunk_start + len(section),
                    }
                    cid = sha1_id(
                        f"{rel_path}|sheet={ws.title}|table|rows={chunk_start+1}-{chunk_start+len(section)}"
                    )
                    chunks.append({"id": cid, "content": text, "metadata": meta})
                    stats["table_chunks"] += 1

            # Embedded images in Excel sheets
            images = list(getattr(ws, "_images", []))
            for img_i, img in enumerate(images, 1):
                try:
                    image_bytes = img._data()
                except Exception:
                    continue
                ext = str(getattr(img, "format", "png")).lower()
                image_path = images_dir / rel_path / f"{ws.title}_img_{img_i:03d}.{ext}"
                save_image(image_bytes, image_path)

                ocr = compute_ocr(image_bytes)
                scanned_by_volume = (
                    len(ocr.text) >= args.ocr_char_threshold or ocr.word_count >= args.ocr_word_threshold
                )
                scanned = scanned_by_volume and (ocr.avg_conf >= args.ocr_min_confidence_for_scanned)
                force_vision = args.image_handling == "vision_only"
                should_use_vision = force_vision or (not scanned)
                content_type = "ocr_text"
                content = ocr.text
                vision_used = False
                vision_error = None
                vision_structured: dict[str, Any] | None = None
                tiled = False
                tile_count = 1
                tile_rows = 1
                tile_cols = 1
                image_width = None
                image_height = None
                image_pixels = None
                vision_retry_used = False
                vision_json_invalid_after_retry = False

                if should_use_vision:
                    stats["images_routed_to_vision"] += 1
                    if vision is None:
                        vision_error = "vision_client_not_ready"
                        if force_vision:
                            fallback = [f"[vision_error] {vision_error}"]
                            if ocr.text:
                                fallback.append(f"OCR fallback: {ocr.text}")
                            content = clean_text(" ".join(fallback)) or "[vision_error_no_fallback_text]"
                            content_type = "image_description"
                        else:
                            content_type = "ocr_text"
                            content = ocr.text or "[vision_failed_and_no_ocr_text]"
                    else:
                        vision_result = describe_image_with_strategy(
                            vision=vision,
                            image_bytes=image_bytes,
                            mime_type=guess_mime(ext),
                            caption_hint=f"Excel sheet {ws.title}",
                            source_file=rel_path,
                            locator=f"sheet {ws.title}, image {img_i}",
                            args=args,
                        )
                        stats["vision_calls_attempted"] += int(vision_result.calls_attempted)
                        stats["vision_calls_succeeded"] += int(vision_result.calls_succeeded)
                        stats["vision_errors"] += int(vision_result.call_errors)
                        stats["vision_input_tokens"] += int(vision_result.input_tokens)
                        stats["vision_output_tokens"] += int(vision_result.output_tokens)
                        stats["vision_total_tokens"] += int(vision_result.total_tokens)

                        tiled = bool(vision_result.tiled)
                        tile_count = int(vision_result.tile_count)
                        tile_rows = int(vision_result.tile_rows)
                        tile_cols = int(vision_result.tile_cols)
                        image_width = int(vision_result.image_width)
                        image_height = int(vision_result.image_height)
                        image_pixels = int(vision_result.image_pixels)
                        vision_retry_used = bool(vision_result.retry_used)
                        vision_json_invalid_after_retry = bool(vision_result.json_invalid_after_retry)
                        if tiled:
                            stats["tiled_images"] += 1
                        if vision_retry_used:
                            stats["vision_retries"] += 1

                        if vision_result.structured is not None:
                            vision_structured = vision_result.structured
                            content = build_image_text_content(vision_structured)
                            content_type = "image_description"
                            vision_used = True
                            if bool(vision_structured.get("json_parse_fallback_used", False)):
                                stats["vision_json_fallbacks"] += 1
                            if vision_result.error:
                                vision_error = vision_result.error
                        else:
                            vision_error = vision_result.error or "vision_strategy_failed_without_output"
                            if force_vision:
                                fallback = [f"[vision_error] {vision_error}"]
                                if ocr.text:
                                    fallback.append(f"OCR fallback: {ocr.text}")
                                content = clean_text(" ".join(fallback)) or "[vision_error_no_fallback_text]"
                                content_type = "image_description"
                            else:
                                content_type = "ocr_text"
                                content = ocr.text or "[vision_failed_and_no_ocr_text]"

                anchor_info = None
                anchor = getattr(img, "anchor", None)
                if anchor is not None and hasattr(anchor, "_from"):
                    anchor_info = {
                        "from_row": int(anchor._from.row) + 1,
                        "from_col": int(anchor._from.col) + 1,
                    }

                meta = {
                    "filename": file_path.name,
                    "source_path": rel_path,
                    "sheet_name": ws.title,
                    "content_type": content_type,
                    "image_index_on_sheet": img_i,
                    "image_path": str(image_path),
                    "image_ext": ext,
                    "anchor": anchor_info,
                    "ocr_word_count": ocr.word_count,
                    "ocr_avg_conf": ocr.avg_conf,
                    "ocr_char_count": len(ocr.text),
                    "scanned_detected_by_volume": scanned_by_volume,
                    "scanned_text_detected": scanned,
                    "image_handling": args.image_handling,
                    "vision_used": vision_used,
                    "vision_provider": args.vision_provider if should_use_vision else None,
                    "vision_model": args.vision_model if should_use_vision else None,
                    "vision_error": vision_error,
                    "tiled": tiled,
                    "tile_count": tile_count,
                    "tile_rows": tile_rows,
                    "tile_cols": tile_cols,
                    "image_width": image_width,
                    "image_height": image_height,
                    "image_pixels": image_pixels,
                    "vision_retry_used": vision_retry_used,
                    "vision_json_invalid_after_retry": vision_json_invalid_after_retry,
                }
                if vision_structured is not None:
                    extracted_text_lines = to_string_list(vision_structured.get("visible_text", []))
                    meta["image_type"] = vision_structured.get("image_type")
                    meta["extracted_text"] = "\n".join(extracted_text_lines)
                    meta["visible_text_lines"] = extracted_text_lines
                    meta["description"] = vision_structured.get("description")
                    meta["elements"] = to_string_list(vision_structured.get("elements", []))
                    meta["layout"] = vision_structured.get("layout")
                    meta["unclear_parts"] = vision_structured.get("unclear_parts")
                    meta["vision_json_parse_fallback_used"] = bool(
                        vision_structured.get("json_parse_fallback_used", False)
                    )
                cid = sha1_id(f"{rel_path}|sheet={ws.title}|image|{img_i}|{content_type}")
                chunks.append({"id": cid, "content": content, "metadata": meta})
                stats["image_chunks"] += 1
                if content_type == "ocr_text":
                    stats["ocr_chunks"] += 1
                if content_type == "image_description":
                    stats["vision_chunks"] += 1
    finally:
        wb.close()

    return chunks, stats


def write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    safe_mkdir(path.parent)
    with path.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=True) + "\n")


def write_per_file_json(
    per_file_root: Path,
    rel_path: str,
    file_type: str,
    stats: dict[str, Any],
    chunks: list[dict[str, Any]],
) -> Path:
    out_path = per_file_root / f"{rel_path}.extraction.json"
    safe_mkdir(out_path.parent)
    counts_by_type: dict[str, int] = {}
    for c in chunks:
        ct = c.get("metadata", {}).get("content_type", "unknown")
        counts_by_type[ct] = counts_by_type.get(ct, 0) + 1
    payload = {
        "source_path": rel_path,
        "file_type": file_type,
        "chunk_count": len(chunks),
        "counts_by_content_type": counts_by_type,
        "stats": stats,
        "chunks": chunks,
    }
    out_path.write_text(json.dumps(payload, indent=2, ensure_ascii=True), encoding="utf-8")
    return out_path


def try_store_chroma(chunks: list[dict[str, Any]], args: argparse.Namespace) -> dict[str, Any]:
    if args.vector_db != "chroma":
        return {"enabled": False}

    try:
        import chromadb
    except Exception as exc:
        return {"enabled": True, "ok": False, "error": f"chromadb import failed: {exc}"}

    safe_mkdir(Path(args.chroma_path))
    try:
        client = chromadb.PersistentClient(path=args.chroma_path)
        collection = client.get_or_create_collection(name=args.chroma_collection)

        ids = [c["id"] for c in chunks]
        docs = [c["content"] for c in chunks]
        metas = []
        for c in chunks:
            m = dict(c["metadata"])
            # Chroma metadata must be flat/scalar values.
            flattened = {}
            for k, v in m.items():
                if v is None:
                    continue
                if isinstance(v, (str, int, float, bool)):
                    flattened[k] = v
                else:
                    flattened[k] = json.dumps(v, ensure_ascii=True)
            metas.append(flattened)

        for start in range(0, len(chunks), args.chroma_batch_size):
            end = min(len(chunks), start + args.chroma_batch_size)
            collection.add(
                ids=ids[start:end],
                documents=docs[start:end],
                metadatas=metas[start:end],
            )

        return {
            "enabled": True,
            "ok": True,
            "collection": args.chroma_collection,
            "path": args.chroma_path,
            "records_added": len(chunks),
        }
    except Exception as exc:
        return {"enabled": True, "ok": False, "error": str(exc)}


def main() -> int:
    args = parse_args()
    data_dir = Path(args.data_dir).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()
    images_dir = output_dir / "extracted_images"
    per_file_json_dir = output_dir / "per_file_json"
    chunks_path = output_dir / "chunks.jsonl"
    summary_path = output_dir / "summary.json"

    if not data_dir.exists() or not data_dir.is_dir():
        raise SystemExit(f"Data directory does not exist or is not a directory: {data_dir}")

    safe_mkdir(output_dir)
    safe_mkdir(images_dir)
    safe_mkdir(per_file_json_dir)

    vision: VisionDescriber | None = None
    vision_init_error = None
    if args.vision_provider != "none":
        try:
            vision = VisionDescriber(args.vision_provider, args.vision_model, args.vision_max_tokens)
        except Exception as exc:
            vision_init_error = str(exc)

    if args.image_handling == "vision_only" and vision is None:
        reason = vision_init_error or "vision provider is none"
        raise SystemExit(
            "image-handling is vision_only, but vision client is not ready. "
            f"Reason: {reason}. Set OPENAI_API_KEY/ANTHROPIC_API_KEY and choose a provider."
        )

    files = list_target_files(data_dir)
    all_chunks: list[dict[str, Any]] = []
    per_file_stats: list[dict[str, Any]] = []
    failed_files: list[dict[str, str]] = []

    for file_path in files:
        rel_path = str(file_path.relative_to(data_dir))
        ext = file_path.suffix.lower()
        fstats = {"source_path": rel_path}

        try:
            if ext == ".pdf":
                chunks, stats = process_pdf(file_path, rel_path, images_dir, args, vision)
                fstats.update({"file_type": "pdf", **stats, "chunk_count": len(chunks)})
                per_file_json_path = write_per_file_json(
                    per_file_root=per_file_json_dir,
                    rel_path=rel_path,
                    file_type="pdf",
                    stats=fstats,
                    chunks=chunks,
                )
                fstats["per_file_json"] = str(per_file_json_path)
                all_chunks.extend(chunks)
            elif ext == ".xlsx":
                chunks, stats = process_excel(file_path, rel_path, images_dir, args, vision)
                fstats.update({"file_type": "xlsx", **stats, "chunk_count": len(chunks)})
                per_file_json_path = write_per_file_json(
                    per_file_root=per_file_json_dir,
                    rel_path=rel_path,
                    file_type="xlsx",
                    stats=fstats,
                    chunks=chunks,
                )
                fstats["per_file_json"] = str(per_file_json_path)
                all_chunks.extend(chunks)
            else:
                continue
            per_file_stats.append(fstats)
        except Exception as exc:
            failed_files.append({"source_path": rel_path, "error": str(exc)})

    write_jsonl(chunks_path, all_chunks)
    chroma_result = try_store_chroma(all_chunks, args)

    counts_by_type: dict[str, int] = {}
    for c in all_chunks:
        ct = c["metadata"].get("content_type", "unknown")
        counts_by_type[ct] = counts_by_type.get(ct, 0) + 1

    total_images_extracted = sum(int(s.get("image_chunks", 0)) for s in per_file_stats)
    total_images_routed_to_vision = sum(int(s.get("images_routed_to_vision", 0)) for s in per_file_stats)
    total_vision_api_calls = sum(int(s.get("vision_calls_attempted", 0)) for s in per_file_stats)
    total_vision_success = sum(int(s.get("vision_calls_succeeded", 0)) for s in per_file_stats)
    total_vision_errors = sum(int(s.get("vision_errors", 0)) for s in per_file_stats)
    total_vision_json_fallbacks = sum(int(s.get("vision_json_fallbacks", 0)) for s in per_file_stats)
    total_vision_retries = sum(int(s.get("vision_retries", 0)) for s in per_file_stats)
    total_tiled_images = sum(int(s.get("tiled_images", 0)) for s in per_file_stats)
    total_vision_input_tokens = sum(int(s.get("vision_input_tokens", 0)) for s in per_file_stats)
    total_vision_output_tokens = sum(int(s.get("vision_output_tokens", 0)) for s in per_file_stats)
    total_vision_tokens = sum(int(s.get("vision_total_tokens", 0)) for s in per_file_stats)
    estimated_api_cost_usd = round(
        (total_vision_input_tokens / 1_000_000.0) * float(args.vision_input_cost_per_1m)
        + (total_vision_output_tokens / 1_000_000.0) * float(args.vision_output_cost_per_1m),
        6,
    )

    vision_error_reasons: dict[str, int] = {}
    for c in all_chunks:
        v_err = c.get("metadata", {}).get("vision_error")
        if v_err:
            vision_error_reasons[v_err] = vision_error_reasons.get(v_err, 0) + 1

    summary = {
        "generated_at_utc": now_utc_iso(),
        "data_dir": str(data_dir),
        "output_dir": str(output_dir),
        "file_count": len(files),
        "processed_file_count": len(per_file_stats),
        "failed_file_count": len(failed_files),
        "chunk_count": len(all_chunks),
        "counts_by_content_type": counts_by_type,
        "vision_provider": args.vision_provider,
        "vision_model": args.vision_model if args.vision_provider != "none" else None,
        "image_handling": args.image_handling,
        "vision_ready": vision is not None,
        "vision_init_error": vision_init_error,
        "vision_usage": {
            "total_images_extracted": total_images_extracted,
            "total_images_sent_to_vision": total_images_routed_to_vision,
            "total_vision_api_calls": total_vision_api_calls,
            "vision_calls_succeeded": total_vision_success,
            "vision_errors": total_vision_errors,
            "vision_json_fallbacks": total_vision_json_fallbacks,
            "vision_retries": total_vision_retries,
            "tiled_images": total_tiled_images,
            "input_tokens": total_vision_input_tokens,
            "output_tokens": total_vision_output_tokens,
            "total_tokens": total_vision_tokens,
            "input_cost_per_1m": float(args.vision_input_cost_per_1m),
            "output_cost_per_1m": float(args.vision_output_cost_per_1m),
            "estimated_api_cost_usd": estimated_api_cost_usd,
            "error_reasons": vision_error_reasons,
        },
        "vector_db": args.vector_db,
        "vector_db_result": chroma_result,
        "failed_files": failed_files,
        "per_file_stats": per_file_stats,
        "outputs": {
            "chunks_jsonl": str(chunks_path),
            "images_dir": str(images_dir),
            "per_file_json_dir": str(per_file_json_dir),
            "summary_json": str(summary_path),
        },
    }
    summary_path.write_text(json.dumps(summary, indent=2, ensure_ascii=True), encoding="utf-8")

    print(f"Processed files: {len(per_file_stats)} / {len(files)}")
    print(f"Generated chunks: {len(all_chunks)}")
    print(f"Chunks JSONL: {chunks_path}")
    print(f"Summary JSON: {summary_path}")
    if args.vector_db == "chroma":
        if chroma_result.get("ok"):
            print(
                f"Chroma stored: {chroma_result.get('records_added')} "
                f"records in {chroma_result.get('collection')}"
            )
        else:
            print(f"Chroma store failed: {chroma_result.get('error')}")
    if vision_init_error:
        print(f"Vision disabled due to init error: {vision_init_error}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

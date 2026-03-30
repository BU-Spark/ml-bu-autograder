from __future__ import annotations

from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

import openpyxl

from core.chunking import clean_text, make_sort_key
from image_utils.ocr import compute_ocr


@dataclass
class TableBlock:
    sheet_name: str
    sheet_index: int
    block_index: int
    row_start: int
    row_end: int
    text: str
    sort_key: str
    document_order: int


@dataclass
class ImageItem:
    sheet_name: str
    sheet_index: int
    image_index: int
    block_index: int
    image_path: str
    ext: str
    anchor: dict[str, int] | None
    ocr_text: str
    ocr_word_count: int
    ocr_avg_conf: float
    sort_key: str
    document_order: int


@dataclass
class ExtractedExcel:
    source_path: str
    file_type: str
    table_blocks: list[TableBlock]
    images: list[ImageItem]
    stats: dict[str, Any]


def _find_header_row(sheet_rows: list[list[str]]) -> int | None:
    best_idx = None
    best_score = float("-inf")
    header_keywords = {
        "header",
        "function",
        "importance",
        "score",
        "notes",
        "description",
        "id",
        "name",
        "requirement",
        "vendor",
    }
    for i, row in enumerate(sheet_rows[:120]):
        non_empty = [v for v in row if v]
        n = len(non_empty)
        if n < 2:
            continue
        lengths = [len(v) for v in non_empty]
        avg_len = sum(lengths) / max(1, n)
        alpha_cells = sum(1 for v in non_empty if any(ch.isalpha() for ch in v))
        keyword_hits = sum(
            1
            for v in non_empty
            for token in [t.lower() for t in v.split()]
            if token in header_keywords
        )
        score = 0.0
        score += n * 1.4
        score += alpha_cells
        score += keyword_hits * 2.5
        if 3 <= avg_len <= 30:
            score += 5.0
        if avg_len > 70:
            score -= 8.0
        if any(len(v) > 130 for v in non_empty):
            score -= 4.0
        if n >= 3:
            score += 2.0
        score -= i * 0.03
        if score > best_score:
            best_score = score
            best_idx = i
    return best_idx


def _save_image(image_bytes: bytes, image_path: Path) -> None:
    image_path.parent.mkdir(parents=True, exist_ok=True)
    image_path.write_bytes(image_bytes)


def extract_excel(file_path: Path, rel_path: str, extract_root: Path, cfg: dict[str, Any]) -> ExtractedExcel:
    wb = openpyxl.load_workbook(file_path, data_only=False)
    table_blocks: list[TableBlock] = []
    images: list[ImageItem] = []
    stats: dict[str, Any] = {
        "table_blocks": 0,
        "images": 0,
    }
    doc_order = 0

    try:
        for sheet_index, ws in enumerate(wb.worksheets, 1):
            max_rows = min(ws.max_row or 1, int(cfg.get("max_sheet_rows", 600)))
            max_cols = min(ws.max_column or 1, int(cfg.get("max_sheet_cols", 60)))

            grid: list[list[str]] = []
            for r in ws.iter_rows(min_row=1, max_row=max_rows, min_col=1, max_col=max_cols, values_only=True):
                grid.append([clean_text("" if v is None else str(v)) for v in r])

            header_idx = _find_header_row(grid)
            header = grid[header_idx] if header_idx is not None else []
            if header:
                header_line = " | ".join([h if h else f"col_{i+1}" for i, h in enumerate(header)])
            else:
                header_line = " | ".join([f"col_{i+1}" for i in range(max_cols)])

            data_lines: list[str] = []
            start_row = (header_idx + 1) if header_idx is not None else 0
            for ridx in range(start_row, len(grid)):
                row = grid[ridx]
                if not any(row):
                    continue
                pairs = []
                for cidx, value in enumerate(row):
                    if not value:
                        continue
                    key = header[cidx] if cidx < len(header) and header[cidx] else f"col_{cidx+1}"
                    pairs.append(f"{key}: {value}")
                if pairs:
                    data_lines.append(f"row_{ridx+1}: " + " ; ".join(pairs))

            rows_per_chunk = int(cfg.get("table_rows_per_chunk", 35))
            for chunk_start in range(0, len(data_lines), rows_per_chunk):
                section = data_lines[chunk_start : chunk_start + rows_per_chunk]
                block_index = chunk_start // rows_per_chunk
                doc_order += 1
                text = (
                    f"Sheet: {ws.title}\n"
                    f"Headers: {header_line}\n"
                    "Rows:\n" + "\n".join(section)
                )
                table_blocks.append(
                    TableBlock(
                        sheet_name=ws.title,
                        sheet_index=sheet_index,
                        block_index=block_index,
                        row_start=chunk_start + 1,
                        row_end=chunk_start + len(section),
                        text=text,
                        sort_key=make_sort_key(sheet_index, block_index),
                        document_order=doc_order,
                    )
                )

            image_block_start = max(1, (len(data_lines) + max(1, rows_per_chunk) - 1) // max(1, rows_per_chunk))
            for img_i, img in enumerate(list(getattr(ws, "_images", [])), 1):
                try:
                    image_bytes = img._data()
                except Exception:
                    continue

                ext = str(getattr(img, "format", "png")).lower()
                image_path = extract_root / "images" / rel_path / f"{ws.title}_img_{img_i:03d}.{ext}"
                _save_image(image_bytes, image_path)
                ocr = compute_ocr(image_bytes)

                anchor_info = None
                anchor = getattr(img, "anchor", None)
                if anchor is not None and hasattr(anchor, "_from"):
                    anchor_info = {
                        "from_row": int(anchor._from.row) + 1,
                        "from_col": int(anchor._from.col) + 1,
                    }

                block_index = image_block_start + (img_i - 1)
                doc_order += 1
                images.append(
                    ImageItem(
                        sheet_name=ws.title,
                        sheet_index=sheet_index,
                        image_index=img_i,
                        block_index=block_index,
                        image_path=str(image_path),
                        ext=ext,
                        anchor=anchor_info,
                        ocr_text=ocr.text,
                        ocr_word_count=ocr.word_count,
                        ocr_avg_conf=ocr.avg_conf,
                        sort_key=make_sort_key(sheet_index, block_index),
                        document_order=doc_order,
                    )
                )

    finally:
        wb.close()

    stats["table_blocks"] = len(table_blocks)
    stats["images"] = len(images)

    return ExtractedExcel(
        source_path=rel_path,
        file_type="xlsx",
        table_blocks=table_blocks,
        images=images,
        stats=stats,
    )


def extracted_excel_to_jsonable(extracted: ExtractedExcel) -> dict[str, Any]:
    return {
        "source_path": extracted.source_path,
        "file_type": extracted.file_type,
        "stats": extracted.stats,
        "table_blocks": [asdict(x) for x in extracted.table_blocks],
        "images": [asdict(x) for x in extracted.images],
    }

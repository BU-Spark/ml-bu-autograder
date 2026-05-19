"""
blueprints/grading.py — /api/grade, /api/grade-batch, /api/describe, /api/grade-existing, /api/grade-quiz-batch
"""
from __future__ import annotations

import io
import json
import re
import sys
import tempfile
from datetime import datetime
from pathlib import Path
from uuid import uuid4

from flask import Blueprint, jsonify, request, send_file

from web.config import (
    OUTPUT_ROOT,
    PROVIDERS,
    PROVIDER_API_KEY_ENV,
    RUBRIC_ALLOWED_EXTS,
    SCRIPTS_DIR,
    SHARED_LECTURE_COLLECTION,
    STUDENT_ALLOWED_EXTS,
)
from web.utils.files import (
    _collect_supporting_files,
    _is_allowed_ext,
    _resolve_project_file,
    _resolve_selected_path,
    _safe,
    _safe_upload_name,
)
from web.utils.pdf_generator import _generate_grade_pdf
from web.utils.pipeline import (
    _adaptive_top_k,
    _cli,
    _embedding_env,
    _run,
    _shared_chroma_dir,
    _shared_chroma_ready,
)

import os

grading_bp = Blueprint("grading", __name__)


def _original_name(path: Path | None, fallback: str = "—") -> str:
    if not path:
        return fallback
    name = path.name
    if len(name) > 9 and name[8] == "_" and all(c in "0123456789abcdef" for c in name[:8]):
        return name[9:]
    return name


@grading_bp.route("/api/grade", methods=["POST"])
def api_grade():
    student = (
        request.files.get("student_pdf")
        or request.files.get("student_submission")
        or request.files.get("student_file")
    )
    if not student or not student.filename:
        return jsonify(success=False, error="Please upload a student submission file."), 400
    if not _is_allowed_ext(student.filename, STUDENT_ALLOWED_EXTS):
        return jsonify(success=False, error="Invalid student file type. Allowed: PDF, PPTX, XLSX."), 400

    describe_provider = request.form.get("describe_provider", "openai")
    describe_model    = request.form.get("describe_model") or PROVIDERS.get(describe_provider, {}).get("model", "gpt-4o-mini")
    grade_provider    = request.form.get("grade_provider", "openai")
    grade_model       = request.form.get("grade_model") or PROVIDERS.get(grade_provider, {}).get("model", "gpt-4o-mini")

    for provider_name in {describe_provider, grade_provider}:
        key_env_name = PROVIDER_API_KEY_ENV.get(provider_name)
        if key_env_name and not os.getenv(key_env_name):
            return jsonify(success=False, error=f"{key_env_name} is not set. Add it to the project .env file."), 400

    run_id   = f"web_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid4().hex[:4]}"
    run_root = OUTPUT_ROOT / run_id
    upload_dir = run_root / "uploads"
    upload_dir.mkdir(parents=True, exist_ok=True)

    original_student_name = student.filename
    student_path = upload_dir / _safe_upload_name(student.filename)
    student.save(str(student_path))

    rubric_path, assignment_path, support_error = _collect_supporting_files(run_root)
    if support_error:
        return jsonify(success=False, error=support_error), 400

    safe_describe_model = _safe(describe_model)
    extract_dir  = run_root / "extract"
    describe_dir = run_root / f"describe_student_{describe_provider}_{safe_describe_model}"
    retrieval_out = run_root / "retrieval.jsonl"

    use_shared_chroma = _shared_chroma_ready()
    chroma_path = _shared_chroma_dir() if use_shared_chroma else run_root / "chroma_db"

    from web.config import DEFAULT_LECTURE_CHUNKS
    steps: list[dict] = []

    # 1. Extract
    code, out = _run(_cli() + [
        "--mode", "extract",
        "--data-dir", str(upload_dir),
        "--output-root", str(OUTPUT_ROOT),
        "--run-id", run_id,
    ])
    steps.append({"step": "Extract PDF", "ok": code == 0, "log": out[-2000:]})
    if code != 0:
        return jsonify(success=False, error="PDF extraction failed.", steps=steps, log=out[-1000:])

    # 2. Describe
    code, out = _run(_cli() + [
        "--mode", "describe",
        "--extract-dir", str(extract_dir),
        "--describe-dir", str(describe_dir),
        "--vision-provider", describe_provider,
        "--vision-model", describe_model,
        "--prompt-version", "verbose_v2",
    ])
    steps.append({"step": "Analyze Content", "ok": code == 0, "log": out[-2000:]})
    if code != 0:
        return jsonify(success=False, error="Content analysis failed.", steps=steps, log=out[-1000:])

    # 3. Index lectures
    has_lectures = DEFAULT_LECTURE_CHUNKS.exists()
    if has_lectures:
        if not use_shared_chroma:
            chroma_dir = _shared_chroma_dir()
            chroma_dir.mkdir(parents=True, exist_ok=True)
            code, out = _run(_cli() + [
                "--mode", "index",
                "--chunks-jsonl", str(DEFAULT_LECTURE_CHUNKS),
                "--output-root", str(OUTPUT_ROOT),
                "--run-id", run_id,
                "--chroma-path", str(chroma_dir),
                "--chroma-collection", SHARED_LECTURE_COLLECTION,
            ], extra_env=_embedding_env())
            steps.append({"step": "Index Lectures", "ok": code == 0, "log": out[-2000:]})
            if code != 0:
                return jsonify(success=False, error="Lecture indexing failed.", steps=steps)
        else:
            steps.append({"step": "Index Lectures", "ok": True, "log": "Reusing shared lecture index."})

        # 4. Retrieve
        code, out = _run(_cli() + [
            "--mode", "retrieve",
            "--chunks-jsonl", str(describe_dir / "chunks.jsonl"),
            "--output-root", str(OUTPUT_ROOT),
            "--run-id", run_id,
            "--chroma-path", str(chroma_path),
            "--chroma-collection", SHARED_LECTURE_COLLECTION,
            "--retrieval-top-k", str(_adaptive_top_k()),
            "--retrieval-out-jsonl", str(retrieval_out),
        ], extra_env=_embedding_env())
        steps.append({"step": "Retrieve Context", "ok": code == 0, "log": out[-2000:]})
        if code != 0:
            return jsonify(success=False, error="Context retrieval failed.", steps=steps)
    else:
        retrieval_out.parent.mkdir(parents=True, exist_ok=True)
        retrieval_out.write_text("")
        steps.append({"step": "Lectures", "ok": True, "log": "No lecture chunks; grading without RAG context."})

    # 5. Grade
    grade_args = _cli() + [
        "--mode", "grade",
        "--chunks-jsonl", str(describe_dir / "chunks.jsonl"),
        "--retrieval-out-jsonl", str(retrieval_out),
        "--output-root", str(OUTPUT_ROOT),
        "--run-id", run_id,
        "--grading-provider", grade_provider,
        "--grading-model", grade_model,
        "--student-path", student_path.name,
    ]
    if assignment_path and assignment_path.exists():
        grade_args += ["--assignment-file", str(assignment_path)]
    if rubric_path and rubric_path.exists():
        grade_args += ["--rubric-file", str(rubric_path)]

    code, out = _run(grade_args)
    steps.append({"step": "Grade Submission", "ok": code == 0, "log": out[-2000:]})
    if code != 0:
        return jsonify(success=False, error="Grading failed.", steps=steps, log=out[-1000:])

    grades_path = run_root / "grading" / "grades.json"
    if not grades_path.exists():
        return jsonify(success=False, error="grades.json not produced.", steps=steps)

    grades = json.loads(grades_path.read_text(encoding="utf-8"))
    grades["_grading_context"] = {
        "student_file":    original_student_name,
        "rubric_file":     _original_name(rubric_path),
        "assignment_file": _original_name(assignment_path),
        "describe_provider": describe_provider,
        "describe_model":    describe_model,
        "grade_provider":    grade_provider,
        "grade_model":       grade_model,
    }

    try:
        pdf_path = _generate_grade_pdf(grades, run_id)
        pdf_filename = pdf_path.name if pdf_path else None
    except Exception:
        pdf_filename = None

    return jsonify(success=True, grades=grades, run_id=run_id, steps=steps, pdf_report=pdf_filename)


@grading_bp.route("/api/grade-batch", methods=["POST"])
def api_grade_batch():
    """Grade a folder of student submissions.

    Extract → Describe → Index → Retrieve runs once for the whole batch.
    Grading runs per-student so each gets an individual result.
    """
    student_files = request.files.getlist("student_files")
    student_files = [f for f in student_files if f and f.filename]
    if not student_files:
        return jsonify(success=False, error="No student files uploaded."), 400

    invalid = [f.filename for f in student_files if not _is_allowed_ext(f.filename, STUDENT_ALLOWED_EXTS)]
    if invalid:
        return jsonify(success=False,
                       error=f"Invalid file type(s): {', '.join(invalid)}. Allowed: PDF, PPTX, XLSX."), 400

    describe_provider = request.form.get("describe_provider", "openai")
    describe_model    = request.form.get("describe_model") or PROVIDERS.get(describe_provider, {}).get("model", "gpt-4o-mini")
    grade_provider    = request.form.get("grade_provider", "openai")
    grade_model       = request.form.get("grade_model") or PROVIDERS.get(grade_provider, {}).get("model", "gpt-4o-mini")

    for provider_name in {describe_provider, grade_provider}:
        key_env_name = PROVIDER_API_KEY_ENV.get(provider_name)
        if key_env_name and not os.getenv(key_env_name):
            return jsonify(success=False, error=f"{key_env_name} is not set."), 400

    run_id     = f"web_batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid4().hex[:4]}"
    run_root   = OUTPUT_ROOT / run_id
    upload_dir = run_root / "uploads"
    upload_dir.mkdir(parents=True, exist_ok=True)

    # Save all student files, track original → saved name mapping
    saved: list[tuple[str, Path]] = []   # [(original_filename, saved_path), ...]
    for f in student_files:
        saved_name  = _safe_upload_name(f.filename)
        saved_path  = upload_dir / saved_name
        f.save(str(saved_path))
        saved.append((f.filename, saved_path))

    rubric_path, assignment_path, support_error = _collect_supporting_files(run_root)
    if support_error:
        return jsonify(success=False, error=support_error), 400

    safe_describe_model = _safe(describe_model)
    extract_dir   = run_root / "extract"
    describe_dir  = run_root / f"describe_student_{describe_provider}_{safe_describe_model}"
    retrieval_out = run_root / "retrieval.jsonl"

    use_shared_chroma = _shared_chroma_ready()
    chroma_path = _shared_chroma_dir() if use_shared_chroma else run_root / "chroma_db"

    from web.config import DEFAULT_LECTURE_CHUNKS
    pipeline_steps: list[dict] = []

    # ── 1. Extract all files at once ──────────────────────────────────────────
    code, out = _run(_cli() + [
        "--mode", "extract",
        "--data-dir", str(upload_dir),
        "--output-root", str(OUTPUT_ROOT),
        "--run-id", run_id,
    ])
    pipeline_steps.append({"step": "Extract All Submissions", "ok": code == 0, "log": out[-2000:]})
    if code != 0:
        return jsonify(success=False, error="Extraction failed.", steps=pipeline_steps, log=out[-1000:])

    # ── 2. Describe all files at once ─────────────────────────────────────────
    code, out = _run(_cli() + [
        "--mode", "describe",
        "--extract-dir", str(extract_dir),
        "--describe-dir", str(describe_dir),
        "--vision-provider", describe_provider,
        "--vision-model", describe_model,
        "--prompt-version", "verbose_v2",
    ])
    pipeline_steps.append({"step": "Analyze All Submissions", "ok": code == 0, "log": out[-2000:]})
    if code != 0:
        return jsonify(success=False, error="Content analysis failed.", steps=pipeline_steps, log=out[-1000:])

    # ── 3. Index lectures (once) ──────────────────────────────────────────────
    has_lectures = DEFAULT_LECTURE_CHUNKS.exists()
    if has_lectures:
        if not use_shared_chroma:
            chroma_dir = _shared_chroma_dir()
            chroma_dir.mkdir(parents=True, exist_ok=True)
            code, out = _run(_cli() + [
                "--mode", "index",
                "--chunks-jsonl", str(DEFAULT_LECTURE_CHUNKS),
                "--output-root", str(OUTPUT_ROOT),
                "--run-id", run_id,
                "--chroma-path", str(chroma_dir),
                "--chroma-collection", SHARED_LECTURE_COLLECTION,
            ], extra_env=_embedding_env())
            pipeline_steps.append({"step": "Index Lectures", "ok": code == 0, "log": out[-2000:]})
            if code != 0:
                return jsonify(success=False, error="Lecture indexing failed.", steps=pipeline_steps)
        else:
            pipeline_steps.append({"step": "Index Lectures", "ok": True, "log": "Reusing shared lecture index."})

        # ── 4. Retrieve (once for all students) ───────────────────────────────
        code, out = _run(_cli() + [
            "--mode", "retrieve",
            "--chunks-jsonl", str(describe_dir / "chunks.jsonl"),
            "--output-root", str(OUTPUT_ROOT),
            "--run-id", run_id,
            "--chroma-path", str(chroma_path),
            "--chroma-collection", SHARED_LECTURE_COLLECTION,
            "--retrieval-top-k", str(_adaptive_top_k()),
            "--retrieval-out-jsonl", str(retrieval_out),
        ], extra_env=_embedding_env())
        pipeline_steps.append({"step": "Retrieve Context", "ok": code == 0, "log": out[-2000:]})
        if code != 0:
            return jsonify(success=False, error="Context retrieval failed.", steps=pipeline_steps)
    else:
        retrieval_out.parent.mkdir(parents=True, exist_ok=True)
        retrieval_out.write_text("")
        pipeline_steps.append({"step": "Lectures", "ok": True, "log": "No lecture chunks; grading without RAG context."})

    # ── 5. Grade each student individually ────────────────────────────────────
    student_results: list[dict] = []

    for original_name, student_path in saved:
        grade_args = _cli() + [
            "--mode", "grade",
            "--chunks-jsonl", str(describe_dir / "chunks.jsonl"),
            "--retrieval-out-jsonl", str(retrieval_out),
            "--output-root", str(OUTPUT_ROOT),
            "--run-id", run_id,
            "--grading-provider", grade_provider,
            "--grading-model", grade_model,
            "--student-path", student_path.name,
        ]
        if assignment_path and assignment_path.exists():
            grade_args += ["--assignment-file", str(assignment_path)]
        if rubric_path and rubric_path.exists():
            grade_args += ["--rubric-file", str(rubric_path)]

        code, out = _run(grade_args)

        if code != 0 or not (run_root / "grading" / "grades.json").exists():
            student_results.append({
                "student_file": original_name,
                "success": False,
                "error": "Grading failed.",
                "log": out[-500:],
            })
            continue

        grades = json.loads((run_root / "grading" / "grades.json").read_text(encoding="utf-8"))
        grades["_grading_context"] = {
            "student_file":      original_name,
            "rubric_file":       _original_name(rubric_path),
            "assignment_file":   _original_name(assignment_path),
            "describe_provider": describe_provider,
            "describe_model":    describe_model,
            "grade_provider":    grade_provider,
            "grade_model":       grade_model,
        }

        try:
            pdf_path     = _generate_grade_pdf(grades, f"{run_id}_{student_path.stem}")
            pdf_filename = pdf_path.name if pdf_path else None
        except Exception:
            pdf_filename = None

        student_results.append({
            "student_file": original_name,
            "success":      True,
            "grades":       grades,
            "pdf_report":   pdf_filename,
        })

    total   = len(student_results)
    passed  = sum(1 for r in student_results if r["success"])
    return jsonify(
        success=True,
        run_id=run_id,
        pipeline_steps=pipeline_steps,
        results=student_results,
        summary={"total": total, "graded": passed, "failed": total - passed},
    )


@grading_bp.route("/api/describe", methods=["POST"])
def api_describe():
    student = (
        request.files.get("student_pdf")
        or request.files.get("student_submission")
        or request.files.get("student_file")
    )
    if not student or not student.filename:
        return jsonify(success=False, error="Please upload a student submission file."), 400
    if not _is_allowed_ext(student.filename, STUDENT_ALLOWED_EXTS):
        return jsonify(success=False, error="Invalid student file type. Allowed: PDF, PPTX, XLSX."), 400

    provider = request.form.get("describe_provider", "openai")
    model    = request.form.get("describe_model") or PROVIDERS.get(provider, {}).get("model", "gpt-4o-mini")

    key_env_name = PROVIDER_API_KEY_ENV.get(provider)
    if key_env_name and not os.getenv(key_env_name):
        return jsonify(success=False, error=f"{key_env_name} is not set."), 400

    run_id     = f"web_describe_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid4().hex[:4]}"
    run_root   = OUTPUT_ROOT / run_id
    upload_dir = run_root / "uploads"
    upload_dir.mkdir(parents=True, exist_ok=True)

    student_path = upload_dir / _safe_upload_name(student.filename)
    student.save(str(student_path))

    safe_model   = _safe(model)
    extract_dir  = run_root / "extract"
    describe_dir = run_root / f"describe_student_{provider}_{safe_model}"
    steps: list[dict] = []

    code, out = _run(_cli() + [
        "--mode", "extract",
        "--data-dir", str(upload_dir),
        "--output-root", str(OUTPUT_ROOT),
        "--run-id", run_id,
    ])
    steps.append({"step": "Extract Submission", "ok": code == 0, "log": out[-2000:]})
    if code != 0:
        return jsonify(success=False, error="Submission extraction failed.", steps=steps, log=out[-1000:])

    code, out = _run(_cli() + [
        "--mode", "describe",
        "--extract-dir", str(extract_dir),
        "--describe-dir", str(describe_dir),
        "--vision-provider", provider,
        "--vision-model", model,
        "--prompt-version", "verbose_v2",
    ])
    steps.append({"step": "Describe Submission", "ok": code == 0, "log": out[-2000:]})
    if code != 0:
        return jsonify(success=False, error="Describe failed.", steps=steps, log=out[-1000:])

    summary_path = describe_dir / "summary.json"
    if not summary_path.exists():
        return jsonify(success=False, error="summary.json not produced.", steps=steps), 500

    summary = json.loads(summary_path.read_text(encoding="utf-8"))
    return jsonify(success=True, run_id=run_id, steps=steps, describe={
        "summary":      summary,
        "summary_path": str(summary_path),
        "chunks_jsonl": str(describe_dir / "chunks.jsonl"),
        "describe_dir": str(describe_dir),
    })


@grading_bp.route("/api/grade-existing", methods=["POST"])
def api_grade_existing():
    provider       = request.form.get("provider", "openai")
    model          = request.form.get("model") or PROVIDERS.get(provider, {}).get("model", "gpt-4o-mini")
    student_filter = (request.form.get("existing_student_path") or "").strip()
    chunks_jsonl_raw   = (request.form.get("existing_chunks_jsonl") or "").strip()
    retrieval_jsonl_raw = (request.form.get("existing_retrieval_jsonl") or "").strip()

    if not chunks_jsonl_raw:
        return jsonify(success=False, error="Existing chunks.jsonl path is required."), 400
    if not retrieval_jsonl_raw:
        return jsonify(success=False, error="Existing retrieval.jsonl path is required."), 400
    if not student_filter:
        return jsonify(success=False, error="Student path filter is required."), 400

    chunks_jsonl   = _resolve_project_file(chunks_jsonl_raw,   allowed_exts={".jsonl"})
    retrieval_jsonl = _resolve_project_file(retrieval_jsonl_raw, allowed_exts={".jsonl"})
    if chunks_jsonl is None:
        return jsonify(success=False, error="chunks.jsonl path is invalid or outside the project."), 400
    if retrieval_jsonl is None:
        return jsonify(success=False, error="retrieval.jsonl path is invalid or outside the project."), 400

    key_env_name = PROVIDER_API_KEY_ENV.get(provider)
    if key_env_name and not os.getenv(key_env_name):
        return jsonify(success=False, error=f"{key_env_name} is not set."), 400

    run_id   = f"web_gradeexisting_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid4().hex[:4]}"
    run_root = OUTPUT_ROOT / run_id
    run_root.mkdir(parents=True, exist_ok=True)

    rubric_path, assignment_path, support_error = _collect_supporting_files(run_root)
    if support_error:
        return jsonify(success=False, error=support_error), 400

    grade_args = _cli() + [
        "--mode", "grade",
        "--chunks-jsonl", str(chunks_jsonl),
        "--retrieval-out-jsonl", str(retrieval_jsonl),
        "--output-root", str(OUTPUT_ROOT),
        "--run-id", run_id,
        "--grading-provider", provider,
        "--grading-model", model,
        "--student-path", student_filter,
    ]
    if assignment_path and assignment_path.exists():
        grade_args += ["--assignment-file", str(assignment_path)]
    if rubric_path and rubric_path.exists():
        grade_args += ["--rubric-file", str(rubric_path)]

    code, out = _run(grade_args)
    steps = [{"step": "Grade Existing Describe Output", "ok": code == 0, "log": out[-2000:]}]
    if code != 0:
        return jsonify(success=False, error="Grade-only run failed.", steps=steps, log=out[-1000:]), 500

    grades_path = run_root / "grading" / "grades.json"
    if not grades_path.exists():
        return jsonify(success=False, error="grades.json not produced.", steps=steps), 500

    grades = json.loads(grades_path.read_text(encoding="utf-8"))
    grades["_grading_context"] = {
        "student_file":    student_filter,
        "rubric_file":     rubric_path.name if rubric_path else "—",
        "assignment_file": assignment_path.name if assignment_path else "—",
        "describe_provider": "—",
        "describe_model":    "—",
        "grade_provider":    provider,
        "grade_model":       model,
    }
    return jsonify(success=True, grades=grades, run_id=run_id, steps=steps)


@grading_bp.route("/api/grade-quiz-batch", methods=["POST"])
def api_grade_quiz_batch():
    if str(SCRIPTS_DIR) not in sys.path:
        sys.path.insert(0, str(SCRIPTS_DIR))

    from grading.grade_submission import (
        SYSTEM_PROMPT, GRADING_PROVIDERS, DEFAULT_GRADING_MODELS, get_api_key,
        extract_rubric_criteria_from_docx, extract_rubric_criteria,
        ai_extract_rubric_criteria, DEFAULT_RUBRIC_CRITERIA,
        read_text_file, _snap_to_grade_band,
    )

    from web.config import LIBRARY_ASSIGNMENTS_DIR, LIBRARY_QUIZZES_DIR, LIBRARY_RUBRICS_DIR, DEFAULT_RUBRIC_DIR

    xlsx_f = request.files.get("quiz_xlsx")
    if not xlsx_f or not xlsx_f.filename:
        return jsonify(success=False, error="No Excel file uploaded."), 400

    provider      = request.form.get("provider", "anthropic")
    model_req     = (request.form.get("model") or "").strip()
    quiz_question = (request.form.get("quiz_question") or "").strip()
    selected_rubric = (request.form.get("selected_rubric") or "").strip()

    if provider not in GRADING_PROVIDERS:
        return jsonify(success=False, error=f"Unknown provider: {provider}"), 400

    key_name, call_fn = GRADING_PROVIDERS[provider]
    api_key = get_api_key(key_name)
    if not api_key:
        return jsonify(success=False, error=f"{key_name.upper()}_API_KEY not set."), 400
    model = model_req or DEFAULT_GRADING_MODELS.get(provider, "")

    rubric_path: Path | None = None
    rubric_f = request.files.get("rubric")
    if rubric_f and rubric_f.filename:
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=Path(rubric_f.filename).suffix)
        rubric_f.save(tmp.name)
        rubric_path = Path(tmp.name)
    elif selected_rubric:
        rubric_path = _resolve_selected_path(
            selected_rubric,
            allowed_roots=[LIBRARY_RUBRICS_DIR, DEFAULT_RUBRIC_DIR, LIBRARY_ASSIGNMENTS_DIR, LIBRARY_QUIZZES_DIR],
            allowed_exts=RUBRIC_ALLOWED_EXTS,
        )

    rubric_text = read_text_file(rubric_path) if rubric_path else ""
    rubric_criteria: list[dict] = []

    if rubric_path and rubric_path.suffix.lower() == ".json":
        try:
            data = json.loads(rubric_path.read_text(encoding="utf-8"))
            raw  = data.get("criteria", [])
            rubric_criteria = [
                {"criterion_id": f"C{i+1}", "criterion_name": c.get("criterion_name", f"C{i+1}"),
                 "max_points": float(c.get("max_points", 0)), "checklist_items": c.get("checklist_items", [])}
                for i, c in enumerate(raw) if float(c.get("max_points", 0)) > 0
            ]
        except Exception:
            pass
    if not rubric_criteria and rubric_path and rubric_path.suffix.lower() == ".docx":
        rubric_criteria = extract_rubric_criteria_from_docx(rubric_path)
    if not rubric_criteria:
        rubric_criteria = extract_rubric_criteria(rubric_text)
    if not rubric_criteria and rubric_text:
        rubric_criteria = ai_extract_rubric_criteria(rubric_text, api_key, provider)
    if not rubric_criteria:
        rubric_criteria = list(DEFAULT_RUBRIC_CRITERIA)

    total_max    = sum(c["max_points"] for c in rubric_criteria)
    criteria_str = json.dumps(rubric_criteria, indent=2)

    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_f)
        ws = wb.active
    except Exception as e:
        return jsonify(success=False, error=f"Failed to read Excel: {e}"), 400

    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        h = str(ws.cell(1, col).value or "").strip().lower()
        headers[h] = col

    def _find_col(*keywords):
        for kw in keywords:
            for h, c in headers.items():
                if kw in h:
                    return c
        return None

    student_col = _find_col("student number", "student id", "student num", "student", "id") or 1
    answer_col  = _find_col("student answer", "answer", "response", "submission") or 2

    score_col    = ws.max_column + 1
    feedback_col = score_col + 1
    ws.cell(1, score_col).value    = "AI Score"
    ws.cell(1, feedback_col).value = "AI Feedback"

    from openpyxl.styles import PatternFill, Font
    teal_fill  = PatternFill("solid", fgColor="0891B2")
    white_bold = Font(bold=True, color="FFFFFF")
    for col in (score_col, feedback_col):
        ws.cell(1, col).fill = teal_fill
        ws.cell(1, col).font = white_bold

    graded = 0
    errors: list[str] = []
    for row in range(2, ws.max_row + 1):
        answer = str(ws.cell(row, answer_col).value or "").strip()
        if not answer:
            continue
        user_msg = (
            f"Grade the following student quiz answer using the rubric criteria.\n\n"
            f"RUBRIC CRITERIA (JSON):\n{criteria_str}\n\n"
            f"RUBRIC TEXT:\n{rubric_text or '(Use criteria above)'}\n\n"
            f"QUIZ QUESTION:\n{quiz_question or '(Grade based on rubric criteria above)'}\n\n"
            f"STUDENT ANSWER:\n{answer[:3000]}\n\n"
            f"Return standard JSON grading schema."
        )
        try:
            resp   = call_fn(model=model, api_key=api_key, system=SYSTEM_PROMPT, user=user_msg)
            result = resp.get("result", {})
            score    = 0.0
            feedback = ""
            if isinstance(result, dict):
                crit_raw = (result.get("criterion_scores") or result.get("criteria_scores")
                            or result.get("criteria") or [])
                if isinstance(crit_raw, list) and crit_raw:
                    for item in crit_raw:
                        if not isinstance(item, dict):
                            continue
                        awarded = float(item.get("awarded_points", 0) or 0)
                        pct     = float(item.get("checklist_pct", 0) or 0)
                        mp      = float(item.get("max_points", 0) or 0)
                        if not awarded and pct and mp:
                            awarded = _snap_to_grade_band(pct, mp)
                        score += awarded
                else:
                    score = float(result.get("overall_score", 0) or 0)
                feedback = str(result.get("overall_feedback", "") or "").strip()
            score = round(min(total_max, max(0.0, score)), 1)
            ws.cell(row, score_col).value    = score
            ws.cell(row, feedback_col).value = feedback
            graded += 1
        except Exception as exc:
            ws.cell(row, score_col).value    = "ERROR"
            ws.cell(row, feedback_col).value = str(exc)[:120]
            errors.append(f"Row {row}: {exc}")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    orig_stem = re.sub(r"[^\w\-]", "_", Path(xlsx_f.filename).stem)[:40]
    out_name  = f"{orig_stem}_AI_Graded.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=out_name,
    )
